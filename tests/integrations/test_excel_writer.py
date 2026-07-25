"""Tests for the Excel row-matching and upsert logic (scenario #10).

The row-lookup helpers are exercised against in-memory tables, and ``save_tasks``
is exercised end-to-end through a real workbook saved in ``tmp_path``: an existing
id is updated in place, a new id is appended and receives a verbatim copy of the
formula columns from the template row.
"""

from datetime import date

import openpyxl

from sop_pipeline.integrations.excel_table_helpers import (
    find_row,
    link_exists,
    next_row,
    expand_table,
)
from sop_pipeline.integrations.excel_writer import ExcelWriter
from sop_pipeline.models.schemas import Priority, Task, TaskType

# Column indices in the BASE_TAREFAS fixture layout (1-based).
COL_ID = 1
COL_TITULO = 2
COL_DIAS_RESTANTES = 13
COL_ATRASADO = 14
COL_STATUS_PRAZO = 15
FORMULA_COLS = [COL_DIAS_RESTANTES, COL_ATRASADO, COL_STATUS_PRAZO]


def _task(task_id: str, title: str) -> Task:
    """A valid Task for the writer tests."""
    return Task(
        task_id=task_id,
        title=title,
        assignee="Alice Silva",
        priority=Priority.HIGH,
        status="In Progress",
        area="TI",
        creation_date=date(2026, 1, 1),
        due_date=date(2026, 2, 1),
        task_type=TaskType.TASK,
        creator="Carol Lima",
        update_date=date(2026, 1, 5),
    )


# --- row-lookup helpers ----------------------------------------------------------


def test_find_row_returns_row_for_existing_id(simple_id_table) -> None:
    """A present id resolves to its 1-based row number."""
    worksheet, table = simple_id_table
    assert find_row(worksheet, "ROW-2", table) == 3


def test_find_row_returns_none_for_missing_id(simple_id_table) -> None:
    """An absent id resolves to None."""
    worksheet, table = simple_id_table
    assert find_row(worksheet, "NOPE", table) is None


def test_next_row_points_past_last_row(simple_id_table) -> None:
    """next_row returns the first free row after the table."""
    _, table = simple_id_table
    assert next_row(table) == 4


def test_expand_table_grows_the_reference(simple_id_table) -> None:
    """Expanding the table stretches its ref to include the new last row."""
    _, table = simple_id_table
    expand_table(table, 4)
    assert table.ref.endswith(":B4")


def test_link_exists_detects_present_and_absent_pairs(links_table) -> None:
    """link_exists is True only for a recorded task/tag pair."""
    worksheet, table = links_table
    assert link_exists(worksheet, table, "ABC-1", 5) is True
    assert link_exists(worksheet, table, "ABC-1", 6) is False


# --- save_tasks round-trip -------------------------------------------------------


def test_save_tasks_updates_existing_row_in_place(tasks_workbook_path: str) -> None:
    """An existing id overwrites its row without appending a new one."""
    ExcelWriter.save_tasks(tasks_workbook_path, [_task("ABC-1", "Updated title")])

    workbook = openpyxl.load_workbook(tasks_workbook_path)
    worksheet = workbook["BASE_TAREFAS"]
    assert worksheet.cell(row=2, column=COL_ID).value == "ABC-1"
    assert worksheet.cell(row=2, column=COL_TITULO).value == "Updated title"
    # No row appended: the table still ends at row 2.
    assert worksheet.tables["base_tarefas"].ref.endswith("2")
    assert worksheet.cell(row=3, column=COL_ID).value is None


def test_save_tasks_appends_new_row_and_copies_formulas(tasks_workbook_path: str) -> None:
    """A new id is appended and the formula columns are copied from row 2."""
    ExcelWriter.save_tasks(tasks_workbook_path, [_task("ABC-2", "Second task")])

    workbook = openpyxl.load_workbook(tasks_workbook_path)
    worksheet = workbook["BASE_TAREFAS"]

    assert worksheet.cell(row=3, column=COL_ID).value == "ABC-2"
    assert worksheet.cell(row=3, column=COL_TITULO).value == "Second task"
    for column in FORMULA_COLS:
        template = worksheet.cell(row=2, column=column).value
        appended = worksheet.cell(row=3, column=column).value
        assert appended == template
        assert isinstance(appended, str) and appended.startswith("=")
