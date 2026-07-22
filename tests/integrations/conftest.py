"""In-memory openpyxl workbook fixtures for the Excel reader/writer tests.

openpyxl builds workbooks purely in memory, which is more reliable than mocking
its intricate table API. The round-trip fixtures save a minimal but real workbook
to ``tmp_path`` so ``ExcelWriter``/``ExcelReader`` can ``load_workbook`` and
``save`` against a genuine file (inside the temp dir, never the project tree).
"""

from datetime import date

import openpyxl
import pytest
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

# Header layout of BASE_TAREFAS: the 12 mapped columns first (column 1 is the
# upsert id), then the 3 formula columns Python only copies, never computes.
TASK_HEADERS = [
    "id",
    "titulo",
    "responsavel",
    "area",
    "prioridade",
    "status",
    "data_criacao",
    "prazo",
    "data_conclusao",
    "tipo",
    "criador",
    "data_atualizacao",
    "dias_restantes",
    "atrasado",
    "status_prazo",
]

# Formula text kept verbatim in the template row (row 2). The exact formulas do
# not matter; what matters is that _copy_formulas replicates these strings onto
# newly appended rows.
TEMPLATE_FORMULAS = {
    "dias_restantes": "=MAX([@prazo]-TODAY(),0)",
    "atrasado": '=IF([@prazo]<TODAY(),"SIM","NAO")',
    "status_prazo": '="No prazo"',
}


def _write_headers(worksheet, headers: list[str]) -> None:
    """Fill row 1 with the given headers."""
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column, value=header)


def _add_table(worksheet, name: str, last_column: int, last_row: int) -> None:
    """Attach an openpyxl table spanning the header row through ``last_row``."""
    ref = f"A1:{get_column_letter(last_column)}{last_row}"
    worksheet.add_table(Table(displayName=name, ref=ref))


@pytest.fixture
def tasks_workbook_path(tmp_path) -> str:
    """A saved BASE_TAREFAS workbook with one template task (ABC-1) in row 2."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "BASE_TAREFAS"
    _write_headers(worksheet, TASK_HEADERS)

    # These non-formula values are arbitrary template content; no test asserts on
    # them, so they deliberately differ from the values used elsewhere.
    row_two = {
        "id": "ABC-1",
        "titulo": "Template title",
        "responsavel": "Template Person",
        "area": "SOP",
        "prioridade": "Low",
        "status": "Done",
        "data_criacao": date(2025, 6, 10),
        "prazo": date(2025, 7, 15),
        "data_conclusao": date(2025, 7, 12),
        "tipo": "Story",
        "criador": "Template Creator",
        "data_atualizacao": date(2025, 7, 12),
        **TEMPLATE_FORMULAS,
    }
    for column, header in enumerate(TASK_HEADERS, start=1):
        worksheet.cell(row=2, column=column, value=row_two[header])

    _add_table(worksheet, "base_tarefas", len(TASK_HEADERS), 2)

    path = tmp_path / "tasks.xlsx"
    workbook.save(path)
    return str(path)


@pytest.fixture
def employees_workbook_path(tmp_path) -> str:
    """A saved DIM_FUNCIONARIO workbook, including one row with a blank cell."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "DIM_FUNCIONARIO"
    headers = ["nome", "jira_email", "clockify_email"]
    _write_headers(worksheet, headers)

    worksheet.cell(row=2, column=1, value="Alice Silva")
    worksheet.cell(row=2, column=2, value="alice.jira@example.com")
    worksheet.cell(row=2, column=3, value="alice.clockify@example.com")

    # Bob's jira_email cell is intentionally left blank.
    worksheet.cell(row=3, column=1, value="Bob Souza")
    worksheet.cell(row=3, column=3, value="bob.clockify@example.com")

    _add_table(worksheet, "dim_funcionario", len(headers), 3)

    path = tmp_path / "employees.xlsx"
    workbook.save(path)
    return str(path)


@pytest.fixture
def simple_id_table():
    """An in-memory (id, valor) table with two data rows, for row-lookup helpers."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    _write_headers(worksheet, ["id", "valor"])
    worksheet.cell(row=2, column=1, value="ROW-1")
    worksheet.cell(row=3, column=1, value="ROW-2")
    _add_table(worksheet, "simple", 2, 3)
    return worksheet, worksheet.tables["simple"]


@pytest.fixture
def links_table():
    """An in-memory (task_id, tag_id) fact table with one existing association."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    _write_headers(worksheet, ["task_id", "tag_id"])
    worksheet.cell(row=2, column=1, value="ABC-1")
    worksheet.cell(row=2, column=2, value=5)
    _add_table(worksheet, "links", 2, 2)
    return worksheet, worksheet.tables["links"]
