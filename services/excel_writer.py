from datetime import date
from enum import Enum
from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries, get_column_letter
from models.schemas import RegistroHoras, Tarefa, DetalheTarefa


class ExcelWriter:

    @staticmethod
    def _find_line(ws, id_wanted, table) -> int | None:
        
        # Default: All Excel tab names are in uppercase letter
        # It's necessary that already have some information in spreadsheet 

        for line in range(2, range_boundaries(table.ref)[3] + 1):
            id_line = ws.cell(row=line, column=1).value
            if id_line == id_wanted:
                return line
            
        return None

    @staticmethod
    def _copy_formula(ws, linha_nova: int, linha_modelo: int, map_columns: dict) -> None:
        colunas_formula = ["dias_restantes", "atrasado", "status_prazo"]
        for nome_coluna in colunas_formula:
            coluna = map_columns[nome_coluna]
            formula = ws.cell(row=linha_modelo, column=coluna).value
            ws.cell(row=linha_nova, column=coluna, value=formula)
            
    @staticmethod
    def _create_map(ws, table) -> dict:
        map_columns = {}
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        for v in range(min_col, max_col + 1):
            value = ws.cell(row=min_row, column=v).value
            map_columns[value] = v
        return map_columns

    @staticmethod
    def _write_line(ws, line: int, task: Tarefa, map_columns: dict, map_translate: dict) -> None:
        
        # All column names are in lowercase

        for k1, v1 in map_translate.items():
                valor = getattr(task, v1)
                
                if isinstance(valor, Enum):
                    valor = valor.value
                
                colunm = map_columns[k1]
                cell = ws.cell(row=line, column=colunm, value=valor)
                
                if isinstance(valor, date):
                    cell.number_format = "DD/MM/YYYY"
    
    @staticmethod
    def save_tasks(way_file: str, tasks: list[Tarefa]) -> None:
        
        # Default: All Excel tab names are in uppercase letter
        # It's necessary that already have some information in spreadsheet 
        # All column names are in lowercase

        wb = ExcelWriter._abrir_workbook(way_file)
        ws = wb["BASE_TAREFAS"]
        table = ws.tables["base_tarefas"]
        map_columns = ExcelWriter._create_map(ws,table)
        
        # Map translated, left side: excel column name, right side: atribute name
        map_translate = {
            "id": "id",
            "titulo": "title",
            "responsavel": "responsible",
            "area": "area",
            "prioridade": "priority",
            "status": "status",
            "data_criacao": "creation_date",
            "prazo": "term",
            "data_conclusao": "completion_date",
            "tipo": "type",
            "criador": "creator",
            "data_atualizacao": "update_date"
            }

        for task in tasks:
            line = ExcelWriter._find_line(ws, task.id, table)
            if line is not None:
                linha = line
                ExcelWriter._write_line(ws, linha, task, map_columns, map_translate )
            else:
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                linha = max_row + 1 
                letter_min = get_column_letter(min_col)
                letter_max = get_column_letter(max_col)
                novo_ref = f"{letter_min}{min_row}:{letter_max}{linha}"
                table.ref = novo_ref
                ExcelWriter._write_line(ws, linha, task, map_columns, map_translate)
                ExcelWriter._copy_formula(ws, linha, linha_modelo=2, map_columns=map_columns)

        wb.save(way_file)

    @staticmethod
    def _find_label_id(ws, table, nome_procurado: str) -> int | None:

        for line in range(2, range_boundaries(table.ref)[3] + 1):
            name_line = ws.cell(row=line, column=2).value
            if name_line == nome_procurado:
                return ws.cell(row=line, column=1).value
            
        return None
    @staticmethod
    def _next_label_id(ws, table) -> int: 
        todos_os_ids_existentes = []
        for line in range(2, range_boundaries(table.ref)[3] + 1):
            cell = ws.cell(row=line, column=1).value
            if cell is not None:
                todos_os_ids_existentes.append(cell)
            else:
                continue

        if not todos_os_ids_existentes:
            return 1
        return max(todos_os_ids_existentes) + 1 
    
    @staticmethod
    def _get_or_create_label_id(ws, table, nome_etiqueta: str) -> int:
        
        id_existente = ExcelWriter._find_label_id(ws, table, nome_etiqueta)
        if id_existente  is not None:
            return id_existente 
        next_id = ExcelWriter._next_label_id(ws, table)
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        linha = max_row + 1 
        ws.cell(row=linha, column=1, value=next_id)
        ws.cell(row=linha, column=2, value=nome_etiqueta)
        letter_min = get_column_letter(min_col)
        letter_max = get_column_letter(max_col)
        novo_ref = f"{letter_min}{min_row}:{letter_max}{linha}"
        table.ref = novo_ref          
        
        return next_id

    @staticmethod
    def _link_exists(ws, table, id_tarefa: str, id_etiqueta: int) -> bool:  
        for line in range(2, range_boundaries(table.ref)[3] + 1):
            id_tarefa_linha = ws.cell(row=line, column=1).value
            id_etiqueta_linha = ws.cell(row=line, column=2).value
            if id_tarefa_linha == id_tarefa and id_etiqueta_linha == id_etiqueta:
                return True
        return False

    @staticmethod
    # Update or create more space in table called "FATO_TAREFA_ETIQUETA"
    def _create_link(ws, table, id_tarefa: str, id_etiqueta: int) -> None:
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        linha_nova = max_row + 1

        ws.cell(row=linha_nova, column=1, value=id_tarefa)
        ws.cell(row=linha_nova, column=2, value=id_etiqueta)

        letter_min = get_column_letter(min_col)
        letter_max = get_column_letter(max_col)
        novo_ref = f"{letter_min}{min_row}:{letter_max}{linha_nova}"
        table.ref = novo_ref
    
    @staticmethod
    def save_tags(way_file: str, tasks: list[Tarefa]) -> None:
        wb = ExcelWriter._abrir_workbook(way_file)
        ws_labels = wb["DIM_ETIQUETAS"]
        table_labels = ws_labels.tables["dim_etiquetas"]  
        ws_links = wb["FATO_TAREFA_ETIQUETA"]
        table_links = ws_links.tables["fato_tarefa_etiqueta"]   

        for task in tasks:
            for nome_etiqueta in task.tags:
                id_etiqueta = ExcelWriter._get_or_create_label_id(ws_labels, table_labels, nome_etiqueta)
                if not ExcelWriter._link_exists(ws_links, table_links, task.id, id_etiqueta):
                    ExcelWriter._create_link(ws_links, table_links, task.id, id_etiqueta)

        wb.save(way_file)

    @staticmethod
    def save_details(way_file: str, details: list[DetalheTarefa]) -> None:
        wb = ExcelWriter._abrir_workbook(way_file)
        ws = wb["DETALHES_TAREFA"]
        table = ws.tables["detalhes_tarefa"]

        for detail in details:
            line = ExcelWriter._find_line(ws, detail.id_tarefa, table)
            if line is not None:
                linha = line
                ExcelWriter._write_detail_line(ws, linha, detail)
            else:
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                linha = max_row + 1
                letter_min = get_column_letter(min_col)
                letter_max = get_column_letter(max_col)
                novo_ref = f"{letter_min}{min_row}:{letter_max}{linha}"
                table.ref = novo_ref
                ExcelWriter._write_detail_line(ws, linha, detail)

        wb.save(way_file)
    
    @staticmethod
    def _write_detail_line(ws, line: int, detail: DetalheTarefa) -> None:
        ws.cell(row=line, column=1, value=detail.id_tarefa)
        ws.cell(row=line, column=2, value=detail.descricao)
    
    @staticmethod
    def _write_hour_line(ws, line: int, registro: RegistroHoras) -> None:
        ws.cell(row=line, column=1, value=registro.id)
        ws.cell(row=line, column=2, value=registro.funcionario)
        ws.cell(row=line, column=4, value=registro.horas)

        cell = ws.cell(row=line, column=3, value=registro.data)

        if isinstance(registro.data, date):
            cell.number_format = "DD/MM/YYYY"

    @staticmethod
    def save_hours(way_file: str, registros: list[RegistroHoras]) -> None:

        wb = ExcelWriter._abrir_workbook(way_file)
        ws = wb["BASE_HORAS"]
        table = ws.tables["base_horas"]

        for registro in registros:
            line = ExcelWriter._find_line(ws, registro.id, table)
            if line is not None:
                ExcelWriter._write_hour_line(ws, line, registro) 
            else:
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                linha = max_row + 1
                letter_min = get_column_letter(min_col)
                letter_max = get_column_letter(max_col)
                novo_ref = f"{letter_min}{min_row}:{letter_max}{linha}"
                table.ref = novo_ref
                ExcelWriter._write_hour_line(ws, linha, registro)
        
        wb.save(way_file)

    @staticmethod
    def _abrir_workbook(way_file: str) -> Workbook:
        wb = load_workbook(way_file)
        wb.calculation.fullCalcOnLoad = True
        return wb