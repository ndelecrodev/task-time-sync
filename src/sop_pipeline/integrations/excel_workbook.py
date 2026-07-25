"""Shared helpers for opening the workbook and mapping its table headers."""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import range_boundaries


def open_workbook(file_path: str) -> Workbook:
    """Load the workbook and force a full recalculation on next open.

    The calculated columns live in Excel formulas, and openpyxl writes back
    the cached values it read. Setting ``fullCalcOnLoad`` makes Excel refresh
    every formula the next time a human opens the file.

    Args:
        file_path: Path to the local workbook.

    Returns:
        Workbook: The loaded workbook.
    """
    workbook = load_workbook(file_path)
    workbook.calculation.fullCalcOnLoad = True
    return workbook


def create_column_map(worksheet, table) -> dict:
    """Map each column header to its column index.

    Args:
        worksheet: The worksheet holding the table.
        table: The openpyxl table whose header row is read.

    Returns:
        dict: Header text mapped to the 1-based column index.
    """
    column_map = {}
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    for column in range(min_col, max_col + 1):
        header = worksheet.cell(row=min_row, column=column).value
        column_map[header] = column
    return column_map
