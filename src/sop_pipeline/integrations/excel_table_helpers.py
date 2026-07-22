"""Row-lookup and structural helpers shared by the Excel reader and writer.

These functions locate rows, compute the next free index, or check whether a
row already exists. None of them place a value into a cell, which is why they
live apart from :class:`~sop_pipeline.integrations.excel_writer.ExcelWriter`'s
save_* methods.
"""

from openpyxl.utils import get_column_letter, range_boundaries


def find_row(worksheet, wanted_id, table) -> int | None:
    """Locate the row holding a given ID inside a table.

    Scans the first column, which by convention holds the table's ID. The
    table must already contain at least the header plus one row, since row 2
    also serves as the formula template.

    Args:
        worksheet: The worksheet to scan.
        wanted_id: The ID to look for.
        table: The openpyxl table delimiting the scan range.

    Returns:
        int | None: The 1-based row number, or ``None`` when not found.
    """
    for row in range(2, range_boundaries(table.ref)[3] + 1):
        if worksheet.cell(row=row, column=1).value == wanted_id:
            return row

    return None


def expand_table(table, last_row: int) -> None:
    """Grow a table's reference range so it covers a newly appended row.

    openpyxl does not track appended rows automatically; without this the new
    row would sit outside the table and be invisible to its formulas.

    Args:
        table: The openpyxl table to resize.
        last_row: The new last row of the table.
    """
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    letter_min = get_column_letter(min_col)
    letter_max = get_column_letter(max_col)
    table.ref = f"{letter_min}{min_row}:{letter_max}{last_row}"


def next_row(table) -> int:
    """Return the row index right after the table's current last row.

    Args:
        table: The openpyxl table to measure.

    Returns:
        int: The 1-based index of the first free row.
    """
    return range_boundaries(table.ref)[3] + 1


def find_tag_id(worksheet, table, wanted_name: str) -> int | None:
    """Look up the surrogate ID of a tag by its name.

    Args:
        worksheet: The ``DIM_ETIQUETAS`` worksheet.
        table: The tag dimension table.
        wanted_name: The tag name to look for.

    Returns:
        int | None: The tag ID, or ``None`` when the tag is unknown.
    """
    for row in range(2, range_boundaries(table.ref)[3] + 1):
        if worksheet.cell(row=row, column=2).value == wanted_name:
            return worksheet.cell(row=row, column=1).value

    return None


def next_tag_id(worksheet, table) -> int:
    """Compute the next free surrogate ID in the tag dimension.

    Args:
        worksheet: The ``DIM_ETIQUETAS`` worksheet.
        table: The tag dimension table.

    Returns:
        int: One past the highest existing ID, or 1 when the table is empty.
    """
    existing_ids = []
    for row in range(2, range_boundaries(table.ref)[3] + 1):
        cell_value = worksheet.cell(row=row, column=1).value
        if cell_value is not None:
            existing_ids.append(cell_value)

    if not existing_ids:
        return 1
    return max(existing_ids) + 1


def link_exists(worksheet, table, task_id: str, tag_id: int) -> bool:
    """Check whether a task-tag association is already recorded.

    Args:
        worksheet: The ``FATO_TAREFA_ETIQUETA`` worksheet.
        table: The fact table holding the associations.
        task_id: Jira issue key.
        tag_id: Surrogate tag ID.

    Returns:
        bool: ``True`` when the pair is already present.
    """
    for row in range(2, range_boundaries(table.ref)[3] + 1):
        row_task_id = worksheet.cell(row=row, column=1).value
        row_tag_id = worksheet.cell(row=row, column=2).value
        if row_task_id == task_id and row_tag_id == tag_id:
            return True
    return False
