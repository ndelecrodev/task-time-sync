"""Writes the pipeline output into the Excel workbook.

The workbook is the final deliverable: it is not a dump, it is a living file with
its own tables, formulas and pivot-style sheets. Two conventions come from the
file itself and are relied on throughout this module:

* sheet names are uppercase, table names are lowercase;
* the first column of every table is the ID used to upsert rows.

Every literal sheet name, table name and column header in this module is a real
identifier inside the ``.xlsx`` file. They are deliberately kept in Portuguese —
translating them would break the lookups at runtime.
"""

from datetime import date
from enum import Enum

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from sop_pipeline.errors.exceptions import ExcelWriteError
from sop_pipeline.models.schemas import Task, TaskDetail, TimeEntry

# Left side: literal column header in the spreadsheet.
# Right side: attribute name on the Task model.
TASK_COLUMN_MAP = {
    "id": "task_id",
    "titulo": "title",
    "responsavel": "assignee",
    "area": "area",
    "prioridade": "priority",
    "status": "status",
    "data_criacao": "creation_date",
    "prazo": "due_date",
    "data_conclusao": "completion_date",
    "tipo": "task_type",
    "criador": "creator",
    "data_atualizacao": "update_date",
}

# Columns owned by Excel formulas, not by Python. Kept out of TASK_COLUMN_MAP on
# purpose so the spreadsheet stays able to recalculate them on its own.
FORMULA_COLUMNS = ["dias_restantes", "atrasado", "status_prazo"]


class ExcelWriter:
    """Upserts tasks, tags, details and hours into the workbook."""

    @staticmethod
    def _find_row(worksheet, wanted_id, table) -> int | None:
        """Locate the row holding a given ID inside a table.

        Scans the first column, which by convention holds the table's ID. The
        table must already contain at least the header plus one row, since row 2
        also serves as the formula template.

        Args:
            worksheet: The worksheet to scan.
            wanted_id: The ID to look for.
            table: The openpyxl table delimiting the scan range.

        Returns:
            int | None: The 1-based row number, or ``None`` when not found.
        """
        for row in range(2, range_boundaries(table.ref)[3] + 1):
            if worksheet.cell(row=row, column=1).value == wanted_id:
                return row

        return None

    @staticmethod
    def _copy_formulas(worksheet, new_row: int, template_row: int, column_map: dict) -> None:
        """Replicate the calculated columns onto a newly appended row.

        Copying the formula *text* verbatim works because those formulas use
        structured references such as ``base_tarefas[[#This Row],[prazo]]``, which
        resolve relative to whichever row they sit on. The same string is
        therefore correct on every row, and no offset rewriting is needed.

        Args:
            worksheet: The worksheet being written to.
            new_row: Row that has just been appended.
            template_row: Row to copy the formulas from (row 2 in practice).
            column_map: Header-to-column-index map for this table.
        """
        for column_name in FORMULA_COLUMNS:
            column = column_map[column_name]
            formula = worksheet.cell(row=template_row, column=column).value
            worksheet.cell(row=new_row, column=column, value=formula)

    @staticmethod
    def _create_column_map(worksheet, table) -> dict:
        """Map each column header to its column index.

        Args:
            worksheet: The worksheet holding the table.
            table: The openpyxl table whose header row is read.

        Returns:
            dict: Header text mapped to the 1-based column index.
        """
        column_map = {}
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        for column in range(min_col, max_col + 1):
            header = worksheet.cell(row=min_row, column=column).value
            column_map[header] = column
        return column_map

    @staticmethod
    def _expand_table(table, last_row: int) -> None:
        """Grow a table's reference range so it covers a newly appended row.

        openpyxl does not track appended rows automatically; without this the new
        row would sit outside the table and be invisible to its formulas.

        Args:
            table: The openpyxl table to resize.
            last_row: The new last row of the table.
        """
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        letter_min = get_column_letter(min_col)
        letter_max = get_column_letter(max_col)
        table.ref = f"{letter_min}{min_row}:{letter_max}{last_row}"

    @staticmethod
    def _next_row(table) -> int:
        """Return the row index right after the table's current last row.

        Args:
            table: The openpyxl table to measure.

        Returns:
            int: The 1-based index of the first free row.
        """
        return range_boundaries(table.ref)[3] + 1

    @staticmethod
    def _write_task_row(worksheet, row: int, task: Task, column_map: dict) -> None:
        """Write every mapped attribute of a task onto one row.

        Args:
            worksheet: The worksheet being written to.
            row: Target row number.
            task: The task whose values are written.
            column_map: Header-to-column-index map for this table.
        """
        for header, attribute in TASK_COLUMN_MAP.items():
            value = getattr(task, attribute)

            # Enums must be unwrapped: openpyxl would otherwise write their repr.
            if isinstance(value, Enum):
                value = value.value

            cell = worksheet.cell(row=row, column=column_map[header], value=value)

            if isinstance(value, date):
                cell.number_format = "DD/MM/YYYY"

    @staticmethod
    def save_tasks(file_path: str, tasks: list[Task]) -> None:
        """Upsert tasks into the ``BASE_TAREFAS`` sheet.

        A task whose ID already exists is updated in place; a new one is appended
        and receives a copy of the calculated columns. This makes repeated runs
        idempotent.

        Args:
            file_path: Path to the local workbook.
            tasks: The tasks to persist.

        Raises:
            ExcelWriteError: If the workbook cannot be updated or saved.
        """
        try:
            workbook = ExcelWriter._open_workbook(file_path)
            worksheet = workbook["BASE_TAREFAS"]
            table = worksheet.tables["base_tarefas"]
            column_map = ExcelWriter._create_column_map(worksheet, table)

            for task in tasks:
                row = ExcelWriter._find_row(worksheet, task.task_id, table)
                if row is not None:
                    ExcelWriter._write_task_row(worksheet, row, task, column_map)
                else:
                    row = ExcelWriter._next_row(table)
                    ExcelWriter._expand_table(table, row)
                    ExcelWriter._write_task_row(worksheet, row, task, column_map)
                    ExcelWriter._copy_formulas(
                        worksheet, row, template_row=2, column_map=column_map
                    )

            workbook.save(file_path)
        except (OSError, KeyError, ValueError) as error:
            raise ExcelWriteError(f"Failed to save tasks to {file_path}: {error}") from error

    @staticmethod
    def _find_tag_id(worksheet, table, wanted_name: str) -> int | None:
        """Look up the surrogate ID of a tag by its name.

        Args:
            worksheet: The ``DIM_ETIQUETAS`` worksheet.
            table: The tag dimension table.
            wanted_name: The tag name to look for.

        Returns:
            int | None: The tag ID, or ``None`` when the tag is unknown.
        """
        for row in range(2, range_boundaries(table.ref)[3] + 1):
            if worksheet.cell(row=row, column=2).value == wanted_name:
                return worksheet.cell(row=row, column=1).value

        return None

    @staticmethod
    def _next_tag_id(worksheet, table) -> int:
        """Compute the next free surrogate ID in the tag dimension.

        Args:
            worksheet: The ``DIM_ETIQUETAS`` worksheet.
            table: The tag dimension table.

        Returns:
            int: One past the highest existing ID, or 1 when the table is empty.
        """
        existing_ids = []
        for row in range(2, range_boundaries(table.ref)[3] + 1):
            cell_value = worksheet.cell(row=row, column=1).value
            if cell_value is not None:
                existing_ids.append(cell_value)

        if not existing_ids:
            return 1
        return max(existing_ids) + 1

    @staticmethod
    def _get_or_create_tag_id(worksheet, table, tag_name: str) -> int:
        """Return a tag's ID, creating the dimension row if it does not exist.

        Args:
            worksheet: The ``DIM_ETIQUETAS`` worksheet.
            table: The tag dimension table.
            tag_name: The tag name to resolve.

        Returns:
            int: The existing or newly assigned tag ID.
        """
        existing_id = ExcelWriter._find_tag_id(worksheet, table, tag_name)
        if existing_id is not None:
            return existing_id

        next_id = ExcelWriter._next_tag_id(worksheet, table)
        row = ExcelWriter._next_row(table)
        worksheet.cell(row=row, column=1, value=next_id)
        worksheet.cell(row=row, column=2, value=tag_name)
        ExcelWriter._expand_table(table, row)

        return next_id

    @staticmethod
    def _link_exists(worksheet, table, task_id: str, tag_id: int) -> bool:
        """Check whether a task-tag association is already recorded.

        Args:
            worksheet: The ``FATO_TAREFA_ETIQUETA`` worksheet.
            table: The fact table holding the associations.
            task_id: Jira issue key.
            tag_id: Surrogate tag ID.

        Returns:
            bool: ``True`` when the pair is already present.
        """
        for row in range(2, range_boundaries(table.ref)[3] + 1):
            row_task_id = worksheet.cell(row=row, column=1).value
            row_tag_id = worksheet.cell(row=row, column=2).value
            if row_task_id == task_id and row_tag_id == tag_id:
                return True
        return False

    @staticmethod
    def _create_link(worksheet, table, task_id: str, tag_id: int) -> None:
        """Append a task-tag association to the fact table.

        Args:
            worksheet: The ``FATO_TAREFA_ETIQUETA`` worksheet.
            table: The fact table holding the associations.
            task_id: Jira issue key.
            tag_id: Surrogate tag ID.
        """
        new_row = ExcelWriter._next_row(table)

        worksheet.cell(row=new_row, column=1, value=task_id)
        worksheet.cell(row=new_row, column=2, value=tag_id)

        ExcelWriter._expand_table(table, new_row)

    @staticmethod
    def save_tags(file_path: str, tasks: list[Task]) -> None:
        """Persist task tags as a dimension plus a many-to-many fact table.

        A task can carry several tags and a tag is shared by many tasks, so the
        relation is modelled with ``DIM_ETIQUETAS`` (one row per distinct tag)
        and ``FATO_TAREFA_ETIQUETA`` (one row per association).

        Args:
            file_path: Path to the local workbook.
            tasks: The tasks whose tags are persisted.

        Raises:
            ExcelWriteError: If the workbook cannot be updated or saved.
        """
        try:
            workbook = ExcelWriter._open_workbook(file_path)
            worksheet_tags = workbook["DIM_ETIQUETAS"]
            table_tags = worksheet_tags.tables["dim_etiquetas"]
            worksheet_links = workbook["FATO_TAREFA_ETIQUETA"]
            table_links = worksheet_links.tables["fato_tarefa_etiqueta"]

            for task in tasks:
                for tag_name in task.tags:
                    tag_id = ExcelWriter._get_or_create_tag_id(worksheet_tags, table_tags, tag_name)
                    if not ExcelWriter._link_exists(
                        worksheet_links, table_links, task.task_id, tag_id
                    ):
                        ExcelWriter._create_link(worksheet_links, table_links, task.task_id, tag_id)

            workbook.save(file_path)
        except (OSError, KeyError, ValueError) as error:
            raise ExcelWriteError(f"Failed to save tags to {file_path}: {error}") from error

    @staticmethod
    def save_details(file_path: str, details: list[TaskDetail]) -> None:
        """Upsert task descriptions into the ``DETALHES_TAREFA`` sheet.

        Args:
            file_path: Path to the local workbook.
            details: The task details to persist.

        Raises:
            ExcelWriteError: If the workbook cannot be updated or saved.
        """
        try:
            workbook = ExcelWriter._open_workbook(file_path)
            worksheet = workbook["DETALHES_TAREFA"]
            table = worksheet.tables["detalhes_tarefa"]

            for detail in details:
                row = ExcelWriter._find_row(worksheet, detail.task_id, table)
                if row is None:
                    row = ExcelWriter._next_row(table)
                    ExcelWriter._expand_table(table, row)
                ExcelWriter._write_detail_row(worksheet, row, detail)

            workbook.save(file_path)
        except (OSError, KeyError, ValueError) as error:
            raise ExcelWriteError(f"Failed to save details to {file_path}: {error}") from error

    @staticmethod
    def _write_detail_row(worksheet, row: int, detail: TaskDetail) -> None:
        """Write one task detail onto a row.

        This sheet has a fixed two-column layout (id, descricao), so the columns
        are addressed by position instead of through a header map.

        Args:
            worksheet: The ``DETALHES_TAREFA`` worksheet.
            row: Target row number.
            detail: The detail record to write.
        """
        worksheet.cell(row=row, column=1, value=detail.task_id)
        worksheet.cell(row=row, column=2, value=detail.description)

    @staticmethod
    def _write_hour_row(worksheet, row: int, time_entry: TimeEntry) -> None:
        """Write one time entry onto a row.

        This sheet has a fixed four-column layout (id, funcionario, data, horas),
        so the columns are addressed by position instead of through a header map.

        Args:
            worksheet: The ``BASE_HORAS`` worksheet.
            row: Target row number.
            time_entry: The time entry to write.
        """
        worksheet.cell(row=row, column=1, value=time_entry.entry_id)
        worksheet.cell(row=row, column=2, value=time_entry.employee)
        worksheet.cell(row=row, column=4, value=time_entry.hours)

        cell = worksheet.cell(row=row, column=3, value=time_entry.entry_date)

        if isinstance(time_entry.entry_date, date):
            cell.number_format = "DD/MM/YYYY"

    @staticmethod
    def save_hours(file_path: str, time_entries: list[TimeEntry]) -> None:
        """Upsert Clockify time entries into the ``BASE_HORAS`` sheet.

        Args:
            file_path: Path to the local workbook.
            time_entries: The time entries to persist.

        Raises:
            ExcelWriteError: If the workbook cannot be updated or saved.
        """
        try:
            workbook = ExcelWriter._open_workbook(file_path)
            worksheet = workbook["BASE_HORAS"]
            table = worksheet.tables["base_horas"]

            for time_entry in time_entries:
                row = ExcelWriter._find_row(worksheet, time_entry.entry_id, table)
                if row is None:
                    row = ExcelWriter._next_row(table)
                    ExcelWriter._expand_table(table, row)
                ExcelWriter._write_hour_row(worksheet, row, time_entry)

            workbook.save(file_path)
        except (OSError, KeyError, ValueError) as error:
            raise ExcelWriteError(f"Failed to save hours to {file_path}: {error}") from error

    @staticmethod
    def _open_workbook(file_path: str) -> Workbook:
        """Load the workbook and force a full recalculation on next open.

        The calculated columns live in Excel formulas, and openpyxl writes back
        the cached values it read. Setting ``fullCalcOnLoad`` makes Excel refresh
        every formula the next time a human opens the file.

        Args:
            file_path: Path to the local workbook.

        Returns:
            Workbook: The loaded workbook.
        """
        workbook = load_workbook(file_path)
        workbook.calculation.fullCalcOnLoad = True
        return workbook
